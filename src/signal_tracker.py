import json
import os
import requests
import argparse
from datetime import datetime
import pytz
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from filelock import FileLock
from config import SIGNALS_FILE, KUCOIN_BASE_URL, KUCOIN_TICKER_ENDPOINT
from telegram_sender import send_telegram_message

def load_signals():
    """بارگذاری سیگنال‌ها از فایل JSON"""
    try:
        if os.path.exists(SIGNALS_FILE):
            with open(SIGNALS_FILE, 'r') as f:
                content = f.read()
                signals = json.loads(content) if content.strip() else []
                print(f"Loaded {len(signals)} signals from {SIGNALS_FILE}")
                tehran_tz = pytz.timezone('Asia/Tehran')
                for signal in signals:
                    if 'status' not in signal or signal['status'] not in ['active', 'target_reached', 'stop_loss']:
                        print(f"Fixing invalid status for {signal.get('symbol', 'unknown')}")
                        signal['status'] = 'active'
                    if 'created_at' not in signal:
                        signal['created_at'] = datetime.now(tehran_tz).strftime("%Y-%m-%d %H:%M:%S")
                    # اصلاح زمان‌ها برای افزودن منطقه زمانی
                    try:
                        created_at = datetime.strptime(signal['created_at'], "%Y-%m-%d %H:%M:%S")
                        if created_at.tzinfo is None:
                            signal['created_at'] = tehran_tz.localize(created_at).strftime("%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        signal['created_at'] = datetime.now(tehran_tz).strftime("%Y-%m-%d %H:%M:%S")
                    if 'closed_at' in signal and signal['closed_at']:
                        try:
                            closed_at = datetime.strptime(signal['closed_at'], "%Y-%m-%d %H:%M:%S")
                            if closed_at.tzinfo is None:
                                signal['closed_at'] = tehran_tz.localize(closed_at).strftime("%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            signal['closed_at'] = None
                return signals
        print(f"No signals found at {SIGNALS_FILE}")
        return []
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON from {SIGNALS_FILE}: {e}")
        return []
    except Exception as e:
        print(f"Error loading signals: {e}")
        return []

def save_signals(signals):
    """ذخیره سیگنال‌ها در فایل JSON با قفل فایل"""
    lock = FileLock(f"{SIGNALS_FILE}.lock")
    try:
        with lock:
            os.makedirs(os.path.dirname(SIGNALS_FILE), exist_ok=True)
            with open(SIGNALS_FILE, 'w') as f:
                json.dump(signals, f, indent=2)
            print(f"Saved {len(signals)} signals to {SIGNALS_FILE}")
    except Exception as e:
        print(f"Error saving signals: {e}")
        send_telegram_message(f"❌ خطا در ذخیره سیگنال‌ها: {e}")

def save_signal(signal):
    """ذخیره یک سیگنال جدید"""
    tehran_tz = pytz.timezone('Asia/Tehran')
    if 'entry_price' not in signal:
        signal['entry_price'] = signal.get('current_price')
    if 'status' not in signal:
        signal['status'] = 'active'
    if 'created_at' not in signal:
        signal['created_at'] = datetime.now(tehran_tz).strftime("%Y-%m-%d %H:%M:%S")
    
    signals = load_signals()
    signals.append(signal)
    save_signals(signals)
    print(f"Signal saved: {signal['symbol']} {signal['type']}")

def get_current_price(symbol):
    """دریافت قیمت فعلی از KuCoin"""
    url = f"{KUCOIN_BASE_URL}{KUCOIN_TICKER_ENDPOINT}"
    params = {"symbol": symbol}
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        price = data.get('data', {}).get('price')
        if price:
            print(f"Price for {symbol}: {price}")
            return float(price)
        print(f"No price data for {symbol}: {data}")
        return None
    except Exception as e:
        print(f"Error fetching price for {symbol}: {e}")
        return None

def calculate_profit_loss(signal, current_price):
    """محاسبه درصد سود/زیان"""
    try:
        entry_price = float(signal.get('entry_price', signal['current_price']))
        if signal['status'] in ['target_reached', 'stop_loss']:
            close_price = float(signal.get('closed_price', current_price))
        else:
            close_price = current_price if current_price else entry_price
        if signal['type'] == 'خرید':
            return ((close_price - entry_price) / entry_price) * 100
        else:  # فروش
            return ((entry_price - close_price) / entry_price) * 100
    except (ValueError, TypeError) as e:
        print(f"Error calculating profit/loss for {signal['symbol']}: {e}")
        return None

def calculate_duration(created_at, closed_at):
    """محاسبه مدت‌زمان سیگنال (به ساعت)"""
    tehran_tz = pytz.timezone('Asia/Tehran')
    try:
        created = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
        if created.tzinfo is None:
            created = tehran_tz.localize(created)
        
        if closed_at:
            closed = datetime.strptime(closed_at, "%Y-%m-%d %H:%M:%S")
            if closed.tzinfo is None:
                closed = tehran_tz.localize(closed)
        else:
            closed = datetime.now(tehran_tz)
        
        duration_hours = (closed - created).total_seconds() / 3600
        print(f"Duration calculated: {duration_hours:.2f} hours for created_at={created_at}, closed_at={closed_at}")
        return duration_hours
    except (ValueError, TypeError) as e:
        print(f"Error calculating duration: {e}")
        return None

def update_signal_status():
    """به‌روزرسانی وضعیت سیگنال‌ها"""
    signals = load_signals()
    if not signals:
        print("No signals to update")
        return

    updated = False
    tehran_tz = pytz.timezone('Asia/Tehran')
    for signal in signals:
        if signal['status'] != 'active':
            print(f"Skipping {signal['symbol']}: Already {signal['status']}")
            continue

        current_price = get_current_price(signal['symbol'])
        if current_price is None:
            print(f"Skipping update for {signal['symbol']} due to missing price")
            continue

        try:
            target_price = float(signal['target_price'])
            stop_loss = float(signal['stop_loss'])
        except (ValueError, TypeError) as e:
            print(f"Invalid target_price or stop_loss for {signal['symbol']}: {e}")
            continue

        now_str = datetime.now(tehran_tz).strftime("%Y-%m-%d %H:%M:%S")

        if signal['type'] == 'خرید':
            if current_price >= target_price:
                signal['status'] = 'target_reached'
                signal['closed_price'] = str(current_price)
                signal['closed_at'] = now_str
                updated = True
                print(f"Updated {signal['symbol']}: Target reached at {current_price}")
            elif current_price <= stop_loss:
                signal['status'] = 'stop_loss'
                signal['closed_price'] = str(current_price)
                signal['closed_at'] = now_str
                updated = True
                print(f"Updated {signal['symbol']}: Stop loss hit at {current_price}")
        elif signal['type'] == 'فروش':
            if current_price <= target_price:
                signal['status'] = 'target_reached'
                signal['closed_price'] = str(current_price)
                signal['closed_at'] = now_str
                updated = True
                print(f"Updated {signal['symbol']}: Target reached at {current_price}")
            elif current_price >= stop_loss:
                signal['status'] = 'stop_loss'
                signal['closed_price'] = str(current_price)
                signal['closed_at'] = now_str
                updated = True
                print(f"Updated {signal['symbol']}: Stop loss hit at {current_price}")

    if updated:
        save_signals(signals)
        print("Signals updated successfully")
    else:
        print("No signals were updated")

def send_telegram_file(file_path):
    """ارسال فایل به تلگرام"""
    bot_token = os.environ.get('TELEGRAM_BOT_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')

    if not bot_token or not chat_id:
        print("Error: Telegram credentials not set")
        return False

    if not os.path.exists(file_path):
        print(f"Error: File {file_path} does not exist")
        return False

    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    try:
        with open(file_path, 'rb') as f:
            files = {'document': (os.path.basename(file_path), f)}
            data = {
                'chat_id': chat_id,
                'caption': '📊 گزارش سیگنال‌ها'
            }
            response = requests.post(url, files=files, data=data, timeout=15)
            if response.status_code == 200:
                print(f"File {file_path} sent to Telegram")
                return True
            else:
                print(f"Error sending file to Telegram: {response.text}")
                return False
    except Exception as e:
        print(f"Error sending file to Telegram: {e}")
        return False

def generate_excel_report():
    """تولید گزارش Excel با شیت‌های مختلف"""
    update_signal_status()
    signals = load_signals()
    tehran_tz = pytz.timezone('Asia/Tehran')
    now_str = datetime.now(tehran_tz).strftime("%Y%m%d_%H%M%S")
    output_file = f"data/signals_report_{now_str}.xlsx"

    # آماده‌سازی داده‌ها برای شیت‌ها
    all_signals_data = []
    active_signals_data = []
    for signal in signals:
        current_price = get_current_price(signal['symbol']) if signal['status'] == 'active' else None
        profit_loss = calculate_profit_loss(signal, current_price)
        duration = calculate_duration(signal['created_at'], signal.get('closed_at'))
        
        signal_row = {
            'Symbol': signal['symbol'],
            'Type': signal['type'],
            'Entry_Price': float(signal.get('entry_price', signal['current_price'])),
            'Target_Price': float(signal['target_price']),
            'Stop_Loss': float(signal['stop_loss']),
            'Created_At': signal['created_at'],
            'Status': signal['status'],
            'Closed_Price': float(signal['closed_price']) if signal.get('closed_price') else None,
            'Closed_At': signal.get('closed_at'),
            'Profit_Loss_%': round(profit_loss, 2) if profit_loss is not None else None,
            'Duration_Hours': round(duration, 2) if duration is not None else None,
            'Reasons': signal['reasons'].replace('✅ ', '').replace('\n', '; ')
        }
        all_signals_data.append(signal_row)
        
        if signal['status'] == 'active' and current_price is not None:
            active_signals_data.append({
                'Symbol': signal['symbol'],
                'Type': signal['type'],
                'Entry_Price': float(signal.get('entry_price', signal['current_price'])),
                'Current_Price': current_price,
                'Price_Change_%': round(profit_loss, 2) if profit_loss is not None else None,
                'Created_At': signal['created_at'],
                'Reasons': signal['reasons'].replace('✅ ', '').replace('\n', '; ')
            })

    # محاسبه آمارها
    total_signals = len(signals)
    active_signals = len([s for s in signals if s['status'] == 'active'])
    target_reached = len([s for s in signals if s['status'] == 'target_reached'])
    stop_loss_signals = len([s for s in signals if s['status'] == 'stop_loss'])
    success_rate = (target_reached / (target_reached + stop_loss_signals) * 100) if (target_reached + stop_loss_signals) > 0 else 0
    avg_profit = pd.Series([s['Profit_Loss_%'] for s in all_signals_data if s['Profit_Loss_%'] is not None]).mean()
    avg_duration = pd.Series([s['Duration_Hours'] for s in all_signals_data if s['Duration_Hours'] is not None]).mean()

    stats_data = [
        {'Metric': 'Total Signals', 'Value': total_signals},
        {'Metric': 'Active Signals', 'Value': active_signals},
        {'Metric': 'Target Reached', 'Value': target_reached},
        {'Metric': 'Stop Loss Hit', 'Value': stop_loss_signals},
        {'Metric': 'Success Rate (%)', 'Value': round(success_rate, 2)},
        {'Metric': 'Average Profit/Loss (%)', 'Value': round(avg_profit, 2) if pd.notna(avg_profit) else None},
        {'Metric': 'Average Duration (Hours)', 'Value': round(avg_duration, 2) if pd.notna(avg_duration) else None}
    ]

    # ایجاد فایل Excel
    wb = Workbook()
    
    # شیت 1: تمام سیگنال‌ها
    ws1 = wb.active
    ws1.title = "All Signals"
    headers = ['Symbol', 'Type', 'Entry Price', 'Target Price', 'Stop Loss', 'Created At', 
               'Status', 'Closed Price', 'Closed At', 'Profit/Loss (%)', 'Duration (Hours)', 'Reasons']
    ws1.append(headers)
    for row in all_signals_data:
        ws1.append([row.get(h.replace(' ', '_'), '') for h in headers])

    # شیت 2: سیگنال‌های فعال
    ws2 = wb.create_sheet("Active Signals")
    headers_active = ['Symbol', 'Type', 'Entry Price', 'Current Price', 'Price Change (%)', 'Created At', 'Reasons']
    ws2.append(headers_active)
    for row in active_signals_data:
        ws2.append([row.get(h.replace(' ', '_'), '') for h in headers_active])

    # شیت 3: آمارها
    ws3 = wb.create_sheet("Statistics")
    ws3.append(['Metric', 'Value'])
    for stat in stats_data:
        ws3.append([stat['Metric'], stat['Value']])

    # اعمال استایل به شیت‌ها
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        ws.freeze_panes = ws['A2']

    # ذخیره فایل
    os.makedirs('data', exist_ok=True)
    try:
        wb.save(output_file)
        print(f"Excel report generated: {output_file}")
    except Exception as e:
        print(f"Error saving Excel report: {e}")
        send_telegram_message(f"❌ خطا در تولید فایل Excel: {e}")
        return

    # ارسال اعلان و فایل به تلگرام
    message = (
        f"📊 گزارش سیگنال‌ها تولید شد\n\n"
        f"🟢 سیگنال‌های فعال: {active_signals}\n"
        f"✅ سیگنال‌های موفق: {target_reached}\n"
        f"❌ سیگنال‌های ناموفق: {stop_loss_signals}\n"
        f"📈 نرخ موفقیت: {success_rate:.2f}%\n"
        f"📅 زمان گزارش: {now_str}\n"
        f"📂 فایل: {output_file}"
    )
    if send_telegram_message(message):
        print("Telegram message sent successfully")
    else:
        print("Failed to send Telegram message")

    if send_telegram_file(output_file):
        print("Excel file sent to Telegram successfully")
    else:
        print("Failed to send Excel file to Telegram")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Track and report signal status')
    parser.add_argument('--report', action='store_true', help='Generate and send a status report')
    args = parser.parse_args()

    try:
        if args.report:
            generate_excel_report()
        else:
            update_signal_status()
    except Exception as e:
        print(f"Error in main execution: {e}")
        send_telegram_message(f"❌ خطای سیستمی در گزارش‌دهی: {e}")
